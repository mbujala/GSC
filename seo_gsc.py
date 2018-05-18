"""Program automatycznie pobiera dane z GSC i zapisuje je w pliku xlsx"""

import datetime
import sys
import time

from bs4 import BeautifulSoup
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font, colors, PatternFill
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.common.exceptions import NoSuchElementException

import config

sys.stderr = open("errorlog.txt", "w")


def nowy_klient(skoroszyt, nazwa):
    """Tworzenie arkusza dla nowego klienta w pliku DANE_GSC"""
    white_font = Font(color="ffffff")
    orange_background = PatternFill("solid", start_color="FA4616",
                                    end_color="FA4616")
    wrap_center = Alignment(vertical="center", wrap_text=True)
    skoroszyt.create_sheet(nazwa)
    sheet = skoroszyt[nazwa]

    sheet["A1"] = "Data"
    sheet["B1"] = "Linki prowadzące do witryny"
    sheet["C1"] = "Ręczne działania"
    sheet["D1"] = "Obsługa mobile - błędy"
    sheet["E1"] = "Stan indeksowania"
    sheet["F1"] = "Liczba stron z zablokowanymi zasobami"
    sheet["G1"] = "Liczba błędów indeksowania"
    sheet["H1"] = "Błędy indeksowania szczegóły"
    sheet["I1"] = "Tester robots.txt"
    sheet["J1"] = "Problemy dotyczące bezpieczeństwa"
    sheet["K1"] = "Błędy danych strukturalnych"
    sheet["L1"] = "Błędy hreflang"
    naglowki = ("A1", "B1", "C1", "D1", "E1",
                "F1", "G1", "H1", "I1", "J1", "K1", "L1")
    sheet.freeze_panes = "A2"
    sheet.row_dimensions[1].height = 50
    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 17
    sheet.column_dimensions["C"].width = 9
    sheet.column_dimensions["D"].width = 15
    sheet.column_dimensions["E"].width = 14
    sheet.column_dimensions["F"].width = 16
    sheet.column_dimensions["G"].width = 13
    sheet.column_dimensions["H"].width = 14
    sheet.column_dimensions["I"].width = 21
    sheet.column_dimensions["J"].width = 18
    sheet.column_dimensions["K"].width = 20
    sheet.column_dimensions["L"].width = 13
    for cell in naglowki:
        sheet[cell].font = white_font
        sheet[cell].fill = orange_background
        sheet[cell].alignment = wrap_center


def logon():
    """logowanie do panelu GSC"""
    DRIVER.get(
        "https://www.google.com/webmasters/tools/dashboard?hl=pl&siteUrl="+url)
    time.sleep(2)
    username = DRIVER.find_element_by_id("identifierId")
    username.send_keys(login)
    DRIVER.find_element_by_id("identifierNext").click()
    time.sleep(2)
    password = DRIVER.find_element_by_name("password")
    password.send_keys(haslo)
    DRIVER.find_element_by_id("passwordNext").click()


def logout():
    """wylogowanie z GSC"""
    DRIVER.get("https://www.google.com/webmasters/tools/logout")
    DRIVER.quit()


def external_links(SHEET, ROW_COUNTER):
    """pobieranie liczby backlinków"""
    DRIVER.get("https://www.google.com/webmasters/tools/external-links?hl=pl&siteUrl="+url)
    time.sleep(2)
    backlinks = DRIVER.find_element_by_xpath(
        "//div[@id='backlinks-dashboard']/div/div/table/tbody/tr[2]/td/div").text
    print("#####Linki do Twojej witryny#####")
    print("Liczba linków: "+backlinks)
    if "Brak danych" in backlinks:
        SHEET["B"+str(ROW_COUNTER)] = backlinks
    else:
        SHEET["B"+str(ROW_COUNTER)] = int(backlinks.replace(" ", ""))


def structured_data(SHEET, ROW_COUNTER):
    """Błędy danych uporządkowanych"""
    DRIVER.get("https://www.google.com/webmasters/tools/structured-data?hl=pl&siteUrl="+url)
    time.sleep(2)
    try:
        struct_data = DRIVER.find_element_by_xpath("/html/body/div[1]/div[3]/div[2]/div/div/div/div/div[1]/div/div/div[2]/div[1]/div[2]").text
        print("#####Dane uporządkowane#####")
        print("Elementy z błędami: "+struct_data)
        if ROW_COUNTER > 2:
            try:
                if int(struct_data.replace(" ", "")) > int(SHEET["K"+str(ROW_COUNTER-1)].value):
                    SHEET["K"+str(ROW_COUNTER)] = int(struct_data.replace(" ", ""))
                    SHEET["K"+str(ROW_COUNTER)].font = Font(color=colors.RED)
                else:
                    SHEET["K"+str(ROW_COUNTER)] = int(struct_data.replace(" ", ""))
            except TypeError:
                SHEET["K"+str(ROW_COUNTER)] = int(struct_data.replace(" ", ""))
        else:
            SHEET["K"+str(ROW_COUNTER)] = int(struct_data.replace(" ", ""))
    except NoSuchElementException:
        struct_data = DRIVER.find_element_by_xpath("/html/body/div[1]/div[3]/div[2]/div/div/div/div/div/div[2]/div").text
        if "W Twojej witrynie nie wykryliśmy" in struct_data:
            SHEET["K"+str(ROW_COUNTER)] = "Brak danych strukturalnych w witrynie"
            print(struct_data)
        else:
            print("Błąd")


def hreflang(SHEET, ROW_COUNTER):
    print("#####Hreflang#####")
    DRIVER.get("https://www.google.com/webmasters/tools/i18n?hl=pl&siteUrl="+url)
    time.sleep(2)
    hreflangi = DRIVER.find_element_by_xpath("/html/body/div[1]/div[3]/div[2]/div/div/div[2]/div[2]/div/div[1]").text
    if "Twoja witryna nie zawiera tagów z atrybutem hreflang" in hreflangi:
        print("Brak hreflangów")
        SHEET["L"+str(ROW_COUNTER)] = "Brak"
    else:
        hreflangi = DRIVER.find_element_by_xpath("/html/body/div[1]/div[3]/div[2]/div/div/div[2]/div[2]/div/div[1]/div/div[3]/div[2]").text
        print("Błędy hreflang:" + hreflangi)
        SHEET["L"+str(ROW_COUNTER)] = int(hreflangi.replace(" ", ""))


def manual_action(SHEET, ROW_COUNTER):
    """sprawdzenie recznych kar"""
    DRIVER.get("https://www.google.com/webmasters/tools/manual-action?hl=pl&siteUrl="+url)
    time.sleep(2)
    recznedzialania = DRIVER.find_element_by_xpath(
        "//div[@id='wmx_gwt_feature_MANUAL_ACTION']/div[2]").text
    print("#####Ręczne działania#####")
    if "W witrynie nie znaleziono ręcznych działań" in recznedzialania:
        print(recznedzialania)
        SHEET["C"+str(ROW_COUNTER)] = "Brak"
    else:
        print("Na stronę została nałożona ręczna kara")
        SHEET["C"+str(ROW_COUNTER)] = "Kara"
        SHEET["C"+str(ROW_COUNTER)].font = Font(color=colors.RED)


def mobile_usability(SHEET, ROW_COUNTER):
    """dostosowanie do urzadzen mobilnych"""
    DRIVER.get("https://www.google.com/webmasters/tools/mobile-usability?hl=pl&siteUrl="+url)
    time.sleep(2)
    mobile = DRIVER.find_element_by_xpath("//div[@id='content']/div/div").text
    print("######Obsługa na urządzeniach mobilnych######")
    if "Nie wykryto żadnych problemów z obsługą" in mobile:
        print(mobile)
        SHEET["D"+str(ROW_COUNTER)] = "Brak błędów"
    else:
        bad_mobile_counter = DRIVER.find_element_by_class_name(
            "wmt-legend-count").text
        print("Liczba stron z problemami: "+bad_mobile_counter)
        SHEET["D"+str(ROW_COUNTER)] = int(bad_mobile_counter.replace(" ", ""))


def index_status(SHEET, ROW_COUNTER):
    """Stan indeksowania"""
    DRIVER.get("https://www.google.com/webmasters/tools/index-status?hl=pl&siteUrl="+url)
    time.sleep(2)
    index = DRIVER.find_element_by_xpath(
        "//div[@id='index-status-chart-legend']/div/table/tbody/tr[2]/td/div").text
    print("#####Stan indeksowania#####")
    print("Zaindeksowano łącznie: "+index)
    if ROW_COUNTER > 2:
        try:
            if int(index.replace(" ", "")) < int(SHEET["E"+str(ROW_COUNTER-1)].value):
                SHEET["E"+str(ROW_COUNTER)] = int(index.replace(" ", ""))
                SHEET["E"+str(ROW_COUNTER)].font = Font(color=colors.RED)
            else:
                SHEET["E"+str(ROW_COUNTER)] = int(index.replace(" ", ""))
        except TypeError:
            SHEET["E"+str(ROW_COUNTER)] = int(index.replace(" ", ""))
    else:
        SHEET["E"+str(ROW_COUNTER)] = int(index.replace(" ", ""))


def roboted(SHEET, ROW_COUNTER):
    """zasoby zablokowane dla robotow"""
    DRIVER.get("https://www.google.com/webmasters/tools/roboted?hl=pl&siteUrl="+url)
    time.sleep(2)
    print("#####Zablokowane zasoby#####")
    try:
        blocked = DRIVER.find_element_by_class_name("wmt-legend-label").text
        blocked_counter = DRIVER.find_element_by_xpath(
            "//div[@id='content']/div/div/div/div[2]/div[2]").text
        if "Strony z zablokowanymi zasobami" in blocked:
            print("Liczba stron z zablokowanymi zasobami: "+blocked_counter)
            SHEET["F"+str(ROW_COUNTER)] = int(blocked_counter.replace(" ", ""))
    except NoSuchElementException:
        print("Nie wykryto zablokowanych zasobów")
        SHEET["F"+str(ROW_COUNTER)] = "Brak"


def crawl_errors(SHEET, ROW_COUNTER):
    """lista bledow ineksowania"""
    DRIVER.get("https://www.google.com/webmasters/tools/crawl-errors?hl=pl&siteUrl="+url)
    time.sleep(2)
    print("#####Błędy indeksowania#####")
    url_error = DRIVER.find_element_by_xpath(
        "/html/body/div[1]/div[3]/div[2]/div/div/div/div[3]/div[1]/span/h2/div[2]").text
    if url_error == "W ciągu ostatnich 90 dni nie wykryto błędów. Super!":
        print("Błędy URL-i: W ciągu ostatnich 90 dni nie wykryto błędów.")
        SHEET["G"+str(ROW_COUNTER)] = "Brak"
        SHEET["H"+str(ROW_COUNTER)] = "Brak"
    else:
        page = DRIVER.page_source
        soup = BeautifulSoup(page, "html.parser")
        print("###Wersja na komputery:###")
        desktop = soup.select("div[style='width: 100%; height: 100%; padding: 0px; margin: 0px;']")
        suma_desktop = 0
        table = []
        for d in desktop:
            tab = d.find_all(class_="JX0GPIC-H-b")
            for x in tab:
                temp = x.find(class_="gwt-Label wmt-legend-count").text
                temp = ''.join(temp.split())
                suma_desktop += int(temp)
                table.append(x.find(class_="gwt-Label").text+": "+x.find(
                    class_="gwt-Label wmt-legend-count").text)
                print(x.find(class_="gwt-Label").text+": "+x.find(
                    class_="gwt-Label wmt-legend-count").text)
        desktop_list = ",".join(table)
        print("###Wersja na smartfony:###")
        mobile = soup.select(
            "div[style='width: 100%; height: 100%; padding: 0px; margin: 0px; display: none;']")
        table = []
        suma_mobile = 0
        for m in mobile:
            tab = m.find_all(class_="JX0GPIC-H-b")
            for x in tab:
                temp = x.find(class_="gwt-Label wmt-legend-count").text
                temp = ''.join(temp.split())
                suma_mobile += int(temp)
                table.append(x.find(class_="gwt-Label").text+": "+x.find(
                    class_="gwt-Label wmt-legend-count").text)
                print(x.find(class_="gwt-Label").text+": "+x.find(
                    class_="gwt-Label wmt-legend-count").text)
        mobile_list = ",".join(table)
        suma = suma_desktop + suma_mobile
        if ROW_COUNTER > 2:
            if suma > int(SHEET["G"+str(ROW_COUNTER-1)].value):
                SHEET["G"+str(ROW_COUNTER)] = suma
                SHEET["G"+str(ROW_COUNTER)].font = Font(color=colors.RED)
                SHEET["H"+str(ROW_COUNTER)] = "Desktop: " + desktop_list + " Mobile: " + mobile_list
            else:
                SHEET["G"+str(ROW_COUNTER)] = suma
                SHEET["H"+str(ROW_COUNTER)] = "Desktop: " + desktop_list + " Mobile: " + mobile_list
        else:
            SHEET["G"+str(ROW_COUNTER)] = suma
            SHEET["H"+str(ROW_COUNTER)] = "Desktop: " + desktop_list + " Mobile: " + mobile_list


def robots_testing(SHEET, ROW_COUNTER):
    """sprawdzanie pliku robots.txt"""
    DRIVER.get(
        "https://www.google.com/webmasters/tools/robots-testing-tool?hl=pl&siteUrl="+url)
    time.sleep(2)
    index = DRIVER.find_element_by_xpath("//*[@id=\":a\"]").text
    print("#####Tester pliku robots.txt#####")
    if "Nie można znaleźć pliku robots.txt" in index:
        print("Brak pliku robots.txt!")
        SHEET["I"+str(ROW_COUNTER)] = "Brak!"
        SHEET["I"+str(ROW_COUNTER)].font = Font(color=colors.RED)
    else:
        day = DRIVER.find_element_by_xpath("//*[@id=\":9\"]").text
        error = DRIVER.find_element_by_class_name("robots-error-count").text
        warnings = DRIVER.find_element_by_class_name("robots-warning-count").text
        print("Wersja pliku robots.txt z dnia: "+day[32:42])
        print("Błędy: "+error[7:])
        print("Ostrzeżenia: "+warnings[13:])
        SHEET["I"+str(ROW_COUNTER)] = "Błędy: "+error[7:]+". Ostrzeżenia: "+warnings[13:]


def security_issues(SHEET, ROW_COUNTER):
    """sprawdzanie problemow dotyczacych bezpieczenstwa"""
    DRIVER.get("https://www.google.com/webmasters/tools/security-issues?hl=pl&siteUrl="+url)
    time.sleep(2)
    print("#####Problemy dotyczace bezpieczeństwa#####")
    try:
        index = DRIVER.find_element_by_xpath(
            "/html/body/div[1]/div[3]/div[2]/div/p[2]").text
        print(index[0:58])
        SHEET["J"+str(ROW_COUNTER)] = "Ok"
    except NoSuchElementException:
        print("Wykryto problemy!")
        SHEET["J"+str(ROW_COUNTER)] = "Wykryto problemy!"
        SHEET["J"+str(ROW_COUNTER)].font = Font(color=colors.RED)


def gsc_scraper():
    """pobiera wszystkie dane dla pojedynczego klienta,
    tworzy zakladkę e excelu jeśli konto nie istnieje"""
    if NAZWA_KLIENTA in WB.sheetnames:
        SHEET = WB[NAZWA_KLIENTA]
        ROW_COUNTER = SHEET.max_row + 1
        SHEET["A"+str(ROW_COUNTER)] = NOW
    else:
        nowy_klient(WB, NAZWA_KLIENTA)
        SHEET = WB[NAZWA_KLIENTA]
        ROW_COUNTER = SHEET.max_row + 1
        SHEET["A"+str(ROW_COUNTER)] = NOW
    try:
        logon()
        external_links(SHEET, ROW_COUNTER)
        manual_action(SHEET, ROW_COUNTER)
        mobile_usability(SHEET, ROW_COUNTER)
        index_status(SHEET, ROW_COUNTER)
        roboted(SHEET, ROW_COUNTER)
        crawl_errors(SHEET, ROW_COUNTER)
        robots_testing(SHEET, ROW_COUNTER)
        security_issues(SHEET, ROW_COUNTER)
        structured_data(SHEET, ROW_COUNTER)
        hreflang(SHEET, ROW_COUNTER)
        logout()
        WB.save('Dane_GSC.xlsx')
    except NoSuchElementException:
        print("Program nie mógł znaleźć elementu")

NOW = datetime.datetime.now().strftime("%d-%m-%Y")

WB2 = load_workbook('Klienci.xlsx')
SHEET_KLIENCI = WB2['Klienci']
ROW_COUNTER_KLIENCI = SHEET_KLIENCI.max_row
WB = load_workbook('Dane_GSC.xlsx')

for x in range(2, ROW_COUNTER_KLIENCI+1):
    OPTIONS = Options()
#    OPTIONS.add_argument("--headless")
    DRIVER = webdriver.Firefox(
        firefox_options=OPTIONS, executable_path=r"geckodriver.exe")
    NAZWA_KLIENTA = SHEET_KLIENCI["A"+str(x)].value
    url = SHEET_KLIENCI["B"+str(x)].value
    KONTO = SHEET_KLIENCI["C"+str(x)].value
    if KONTO == "seo":
        login = config.SEO["login"]
        haslo = config.SEO["haslo"]
        gsc_scraper()
    elif KONTO == "seo2":
        login = config.SEO2["login"]
        haslo = config.SEO2["haslo"]
        gsc_scraper()
    elif KONTO == "oferta":
        login = config.OFERTA["login"]
        haslo = config.OFERTA["haslo"]
        gsc_scraper()
    else:
        print("Niepoprawna nazwa konta w pliku Klienci.xlxs")

print('Koniec')
sys.stderr.close()
